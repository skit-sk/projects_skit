"""OFD Abonent — Flask Blueprint."""
import os, json, time
from pathlib import Path
from datetime import datetime, timedelta
from flask import Blueprint, jsonify, render_template, request

_bp_dir = Path(__file__).resolve().parent
_storage_dir = _bp_dir / "storage"

import sys
sys.path.insert(0, str(_bp_dir))

# Lazy load _load_provider from routes.py
def _get_provider(name):
    """Load provider config from routes.py using importlib."""
    import importlib
    _rp = os.path.join(str(_bp_dir), "routes.py")
    _spec = importlib.util.spec_from_file_location("ofd_routes_mod", _rp)
    _mod = importlib.util.module_from_spec(_spec)
    _ss = sys.path.copy()
    sys.path.insert(0, str(_bp_dir))
    _spec.loader.exec_module(_mod)
    sys.path = _ss
    return _mod._load_provider(name)

bp = Blueprint("ofd_abonent", __name__, url_prefix="/ofd_abonent",
               template_folder=str(_bp_dir / "templates"),
               static_folder=str(_bp_dir / "static"),
               static_url_path="/static/ofd_abonent")


@bp.route("/")
def index():
    return render_template("ofd_abonent.html")


@bp.route("/api/list_inns")
def api_list_inns():
    """List all INNs that have data in the storage."""
    from ofd_storage import DATA_ROOT
    if not DATA_ROOT.exists():
        return jsonify([])
    inns = []
    for d in sorted(DATA_ROOT.iterdir()):
        if d.is_dir() and (d / "inn_abonent.json").exists():
            try:
                ab = json.loads((d / "inn_abonent.json").read_text(encoding="utf-8"))
                inns.append({
                    "inn": d.name,
                    "name": ab.get("name", ""),
                    "provider": ab.get("provider", {}).get("name", ""),
                    "kkt_count": len(ab.get("kkt_summary", [])),
                    "last_sync": ab.get("last_sync", ""),
                })
            except Exception:
                inns.append({"inn": d.name, "name": "", "provider": "", "kkt_count": 0, "last_sync": ""})
    return jsonify(inns)


@bp.route("/api/discover_inn", methods=["POST"])
def api_discover_inn():
    """Discover INN from API token using the provider's /v1/inn endpoint."""
    data = request.get_json(force=True, silent=True) or {}
    provider_id = data.get("provider", "yandex_ofd")
    token = data.get("token", "")

    # Import _load_provider from the main routes module
    provider = _get_provider(provider_id)
    if not provider:
        return jsonify({"error": f"Provider '{provider_id}' not found"})

    # Get token with priority: request body → env var
    env_token_name = provider.get("env_token", "")
    token = token or os.environ.get(env_token_name, "")
    if not token:
        return jsonify({"error": "Token not found. Set in .env or pass in request"})

    if provider_id == "yandex_ofd":
        from bot_ofd.yandex_ofd import YandexOfdClient
        client = YandexOfdClient(token=token)
        import urllib.request, json as _json
        url = f"{client.base}/v1/inn"
        try:
            req = urllib.request.Request(url, data=b"{}", headers={"Ofdapitoken": token, "Content-Type": "application/json"}, method="POST")
            with urllib.request.urlopen(req, timeout=15) as r:
                result = _json.loads(r.read())
        except urllib.request.HTTPError as e:
            return jsonify({"error": f"HTTP {e.code}: {e.reason}"})
        except Exception as e:
            return jsonify({"error": str(e)})
        # v1/inn returns either {"inn": "..."} or ["INN"] or [{"inn": "..."}]
        if isinstance(result, list):
            if result and isinstance(result[0], str):
                inn = result[0]
            elif result and isinstance(result[0], dict):
                inn = result[0].get("inn", "") or result[0].get("INN", "")
            else:
                inn = ""
        else:
            inn = result.get("inn", "") or result.get("INN", "")
        if not inn:
            return jsonify({"error": f"INN not found. Response: {str(result)[:200]}"})
        return jsonify({"inn": inn, "provider": provider_id})
    else:
        from bot_ofd.ofd_client import OfdApiClient
        client = OfdApiClient(provider, token=token)
        result = client.call("v1/inn", "POST", {})
        if "error" in result:
            return jsonify({"error": result["error"]})
        if isinstance(result, list):
            inn = result[0] if result and isinstance(result[0], str) else (result[0].get("inn", "") if result and isinstance(result[0], dict) else "")
        else:
            inn = result.get("inn", "") or result.get("INN", "")
        if not inn:
            return jsonify({"error": f"INN not found in response: {str(result)[:200]}"})
        return jsonify({"inn": inn, "provider": provider_id})


@bp.route("/api/save_inn", methods=["POST"])
def api_save_inn():
    """Create skeleton abonent file for a given INN."""
    data = request.get_json(force=True, silent=True) or {}
    inn = data.get("inn", "")
    if not inn or len(str(inn)) < 10:
        return jsonify({"error": "Invalid INN"})

    from ofd_storage import OfdStorage, DATA_ROOT
    provider_id = data.get("provider", "yandex_ofd")
    provider = _get_provider(provider_id) or {}

    base = DATA_ROOT / str(inn)
    base.mkdir(parents=True, exist_ok=True)
    (base / "items").mkdir(exist_ok=True)

    abonent = {
        "inn": str(inn),
        "name": data.get("name", ""),
        "provider": {"id": provider_id, "name": provider.get("provider", provider_id), "env_token": provider.get("env_token", "")},
        "last_sync": "",
        "kkt_summary": [],
    }
    with open(base / "inn_abonent.json", "w", encoding="utf-8") as f:
        json.dump(abonent, f, ensure_ascii=False, indent=2)
    with open(base / "inn_kkt_index.json", "w", encoding="utf-8") as f:
        json.dump({"inn": str(inn), "kkt_list": []}, f, ensure_ascii=False, indent=2)

    # Save token to local .env
    # Priority: explicit token → provider's env var
    token = data.get("token", "")
    if not token:
        env_name = provider.get("env_token", "")
        if env_name:
            token = os.environ.get(env_name, "")
    if token:
        st = OfdStorage(str(inn))
        st.save_token(token)

    return jsonify({"ok": True, "inn": str(inn)})


@bp.route("/api/defaults")
def api_defaults():
    """Return available env tokens for auto-discovery."""
    tokens = []
    for var in ["OFD_YARU_TOKEN", "OFD_OFDRU_TOKEN"]:
        val = os.environ.get(var, "")
        if val:
            prov = "yandex_ofd" if "YARU" in var else "ofd_ru"
            name = "Яндекс.ОФД" if "YARU" in var else "OFD.RU"
            tokens.append({"var": var, "has_token": True, "provider": prov, "name": name})
    return jsonify({"tokens": tokens})


@bp.route("/api/status")
def api_status():
    from ofd_storage import OfdStorage
    inn = request.args.get("inn", "")
    if not inn:
        return jsonify({"inn": None, "has_data": False})
    st = OfdStorage(inn)
    abonent = st.load_abonent()
    has_data = bool(abonent)
    return jsonify({"inn": inn, "has_data": has_data, "abonent": abonent})


@bp.route("/api/delete_inn", methods=["POST"])
def api_delete_inn():
    """Delete all stored data for an INN."""
    from ofd_storage import DATA_ROOT
    import shutil
    data = request.get_json(force=True, silent=True) or {}
    inn = data.get("inn", "")
    if not inn:
        return jsonify({"error": "INN required"})
    path = DATA_ROOT / str(inn)
    if path.exists():
        shutil.rmtree(path)
        return jsonify({"ok": True})
    return jsonify({"error": "INN not found"})


@bp.route("/api/sync", methods=["POST"])
def api_sync():
    from ofd_storage import OfdStorage
    from ofd_sync import sync_abonent

    data = request.get_json(force=True, silent=True) or {}
    inn = data.get("inn", "")
    period_from = data.get("period_from", (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d"))
    period_to = data.get("period_to", datetime.now().strftime("%Y-%m-%d"))

    # Load abonent → get provider → read token from local .env → fallback main .env
    st = OfdStorage(inn)
    abonent = st.load_abonent()
    if not abonent:
        return jsonify({"error": "Abonent not found. Add INN first."})
    stored_provider = abonent.get("provider", {})
    provider_id = stored_provider.get("id", "yandex_ofd")

    provider = _get_provider(provider_id)
    if not provider:
        return jsonify({"error": f"Provider '{provider_id}' not found"})

    token = st.load_token()
    if not token:
        env_token = provider.get("env_token", "")
        token = os.environ.get(env_token, "")

    t0 = time.time()
    result = sync_abonent(inn, provider, token, period_from, period_to, st)
    elapsed = int((time.time() - t0) * 1000)

    if "error" in result:
        return jsonify({"error": result["error"]})
    return jsonify({"ok": True, "elapsed_ms": elapsed, "kkt_count": result.get("kkt_count", 0)})


@bp.route("/api/kkt_index")
def api_kkt_index():
    from ofd_storage import OfdStorage
    inn = request.args.get("inn", "")
    if not inn:
        return jsonify({"error": "INN required"})
    st = OfdStorage(inn)
    return jsonify(st.load_kkt_index())


@bp.route("/api/kkt/<rnm>")
def api_kkt(rnm):
    from ofd_storage import OfdStorage
    inn = request.args.get("inn", "")
    if not inn:
        return jsonify({"error": "INN required"})
    st = OfdStorage(inn)
    return jsonify(st.load_kkt(rnm))


@bp.route("/api/fn/<fn>")
def api_fn(fn):
    from ofd_storage import OfdStorage
    inn = request.args.get("inn", "")
    if not inn:
        return jsonify({"error": "INN required"})
    st = OfdStorage(inn)
    return jsonify(st.load_fn(fn))


@bp.route("/api/fields")
def api_fields():
    """Return all API fields with values from last sync."""
    from ofd_storage import OfdStorage
    inn = request.args.get("inn", "")
    if not inn:
        return jsonify({"error": "INN required"})

    st = OfdStorage(inn)
    fields = []

    # 1. Static fields from provider JSON
    provider = _get_provider("yandex_ofd")
    if provider:
        for ep_name, ep_info in provider.get("endpoints", {}).items():
            method = ep_info.get("method", "POST /")
            desc = ep_info.get("desc", ep_name)
            for pk, pv in ep_info.get("params", {}).items():
                ex = pv.get("example", "") if isinstance(pv, dict) else str(pv)
                fields.append({"api": ep_name, "endpoint": method, "field": pk, "type": type(ex).__name__, "value": str(ex)[:40]})

    # 2. Real data from FN files (shifts)
    fn_files = st.list_fn_files()
    for fp in fn_files:
        fname = fp.stem
        if fname.startswith("fn_") and not fname.endswith("_daily"):
            fn_data = st.load_fn(fname.replace("fn_", ""))
            shifts = fn_data.get("shifts", [])
            if shifts:
                s = shifts[0]
                for fld in ("shift_number", "open_date", "close_date", "receipt_count", "cash_receipts", "card_receipts", "total_sum", "cash_sum", "card_sum"):
                    fields.append({"api": "KKTShift", "endpoint": "POST /v1/KKTShift", "field": f"shifts[].{fld}", "type": type(s.get(fld, "")).__name__, "value": str(s.get(fld, ""))[:40]})

    # 3. Real data from abonent (KKT summary)
    ab = st.load_abonent()
    for k in ab.get("kkt_summary", []):
        fields.append({"api": "KKTbyTradePoint", "endpoint": "POST /v1/KKTbyTradePoint", "field": "kkt.register_number_kkt", "type": "string", "value": k.get("rnm", "")})
        fields.append({"api": "KKTbyTradePoint", "endpoint": "POST /v1/KKTbyTradePoint", "field": "kkt.factory_number_fn", "type": "string", "value": k.get("fn_current", "")})

    return jsonify({"fields": fields, "count": len(fields)})


@bp.route("/api/charts")
def api_charts():
    from ofd_storage import OfdStorage
    inn = request.args.get("inn", "")
    period = request.args.get("period", "day")
    rnm_filter = request.args.getlist("rnm")
    fn_filter = request.args.getlist("fn")
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")

    if not inn:
        return jsonify({"error": "INN required"})

    from datetime import datetime as dt
    f = dt.strptime(date_from, "%Y-%m-%d") if date_from else dt.now() - timedelta(days=30)
    t = dt.strptime(date_to, "%Y-%m-%d") if date_to else dt.now()

    st = OfdStorage(inn)

    # Aggregate from all FN shift files
    shift_agg = {}
    fn_files = st.list_fn_files()
    for fp in fn_files:
        fname = fp.stem
        if fname.startswith("fn_") and not fname.endswith("_daily"):
            fn_key = fname.replace("fn_", "")
            if fn_filter and fn_key not in fn_filter:
                continue
            fn_data = st.load_fn(fn_key)
            rnm = fn_data.get("rnm", "")
            if rnm_filter and rnm not in rnm_filter:
                continue
            for shift in fn_data.get("shifts", []):
                sd = shift.get("open_date", "")[:10]
                if not sd or sd < date_from or sd > date_to:
                    continue
                if sd not in shift_agg:
                    shift_agg[sd] = {"cash": 0, "card": 0, "return": 0, "errors": 0,
                                     "cash_sum": 0, "card_sum": 0, "return_sum": 0}
                shift_agg[sd]["cash"] += int(shift.get("cash_receipts", 0))
                shift_agg[sd]["card"] += int(shift.get("card_receipts", 0))
                shift_agg[sd]["return"] += int(shift.get("return_receipts", 0))
                shift_agg[sd]["cash_sum"] += float(shift.get("cash_sum", 0))
                shift_agg[sd]["card_sum"] += float(shift.get("card_sum", 0))
                shift_agg[sd]["return_sum"] += float(shift.get("return_sum", 0))

    # Build labels and data based on period
    labels, receipts_data, sums_data = [], {"cash": [], "card": [], "return": [], "errors": []}, {"cash": [], "card": [], "return": []}
    delta = (t - f).days + 1

    if period == "day":
        for i in range(delta):
            d = (f + timedelta(days=i)).strftime("%Y-%m-%d")
            labels.append(d)
            ag = shift_agg.get(d, {})
            receipts_data["cash"].append(ag.get("cash", 0))
            receipts_data["card"].append(ag.get("card", 0))
            receipts_data["return"].append(ag.get("return", 0))
            receipts_data["errors"].append(0)
            sums_data["cash"].append(ag.get("cash_sum", 0))
            sums_data["card"].append(ag.get("card_sum", 0))
            sums_data["return"].append(ag.get("return_sum", 0))
    elif period == "week":
        week_agg = {}
        for d, ag in shift_agg.items():
            w = dt.strptime(d, "%Y-%m-%d").strftime("%Y-W%W")
            if w not in week_agg:
                week_agg[w] = {"cash": 0, "card": 0, "return": 0, "errors": 0, "cash_sum": 0, "card_sum": 0, "return_sum": 0}
            for k in ("cash", "card", "return"):
                week_agg[w][k] += ag.get(k, 0)
                week_agg[w][f"{k}_sum"] += ag.get(f"{k}_sum", 0)
        for w in sorted(week_agg.keys()):
            labels.append(w)
            ag = week_agg[w]
            receipts_data["cash"].append(ag["cash"])
            receipts_data["card"].append(ag["card"])
            receipts_data["return"].append(ag["return"])
            receipts_data["errors"].append(0)
            sums_data["cash"].append(ag["cash_sum"])
            sums_data["card"].append(ag["card_sum"])
            sums_data["return"].append(ag["return_sum"])
    elif period == "month":
        month_agg = {}
        for d, ag in shift_agg.items():
            m = d[:7]
            if m not in month_agg:
                month_agg[m] = {"cash": 0, "card": 0, "return": 0, "errors": 0, "cash_sum": 0, "card_sum": 0, "return_sum": 0}
            for k in ("cash", "card", "return"):
                month_agg[m][k] += ag.get(k, 0)
                month_agg[m][f"{k}_sum"] += ag.get(f"{k}_sum", 0)
        for m in sorted(month_agg.keys()):
            labels.append(m)
            ag = month_agg[m]
            receipts_data["cash"].append(ag["cash"])
            receipts_data["card"].append(ag["card"])
            receipts_data["return"].append(ag["return"])
            receipts_data["errors"].append(0)
            sums_data["cash"].append(ag["cash_sum"])
            sums_data["card"].append(ag["card_sum"])
            sums_data["return"].append(ag["return_sum"])

    return jsonify({"labels": labels, "receipts": receipts_data, "sums": sums_data})


@bp.route("/api/export/xls")
def export_xls():
    """Export data to XLS using openpyxl."""
    from ofd_storage import OfdStorage
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    inn = request.args.get("inn", "")
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    aggr = request.args.get("aggr", "0") == "1"
    if not inn:
        return jsonify({"error": "INN required"})

    st = OfdStorage(inn)
    ab = st.load_abonent()

    wb = openpyxl.Workbook()

    # Sheet 1: Сводка
    ws1 = wb.active
    ws1.title = "Сводка"
    ws1.append(["ККТ", "ФН", "Провайдер", "Смены", "Чеков всего", "Нал", "Карта", "Возврат", "Сумма всего", "Сумма нал", "Сумма карта", "Сумма возврат", "Ошибки"])
    for k in ab.get("kkt_summary", []):
        t = k.get("totals", {})
        ws1.append([
            k.get("kkt_name", ""), k.get("fn_current", ""), k.get("provider", ""),
            t.get("shifts", 0),
            t.get("receipts", {}).get("total", 0), t.get("receipts", {}).get("cash", 0),
            t.get("receipts", {}).get("card", 0), t.get("receipts", {}).get("return", 0),
            t.get("sums", {}).get("total", 0), t.get("sums", {}).get("cash", 0),
            t.get("sums", {}).get("card", 0), t.get("sums", {}).get("return", 0),
            t.get("errors", 0),
        ])

    # Sheet 2: Смены
    ws2 = wb.create_sheet("Смены")
    ws2.append(["ФН", "RNM", "Смена", "Дата открытия", "Дата закрытия", "Чеков", "Нал чеков", "Карта чеков", "Сумма", "Нал", "Карта", "Возврат"])
    fn_files = st.list_fn_files()
    for fp in fn_files:
        fname = fp.stem
        if fname.startswith("fn_") and not fname.endswith("_daily"):
            fn_data = st.load_fn(fname.replace("fn_", ""))
            for s in fn_data.get("shifts", []):
                sd = s.get("open_date", "")[:10]
                if sd and sd >= date_from and sd <= date_to:
                    ws2.append([
                        fn_data.get("fn", ""), fn_data.get("rnm", ""),
                        s.get("shift_number"), s.get("open_date", "")[:16], s.get("close_date", "")[:16],
                        s.get("receipt_count"), s.get("cash_receipts"), s.get("card_receipts"),
                        s.get("total_sum"), s.get("cash_sum"), s.get("card_sum"), s.get("return_sum"),
                    ])

    # Sheet 3: Товары
    items_data = st.load_items()
    receipts = items_data.get("receipts", [])
    if receipts:
        ws3 = wb.create_sheet("Товары")
        if aggr:
            ws3.append(["Товар", "Кол-во", "Цена", "Сумма", "Оплата нал", "Оплата карта", "Чеков"])
            from collections import defaultdict
            agg = defaultdict(lambda: {"quantity": 0, "sum": 0.0, "cash_sum": 0.0, "card_sum": 0.0, "receipt_ids": set()})
            for r in receipts:
                if r.get("date", "") < date_from or r.get("date", "") > date_to:
                    continue
                receipt_id = r.get("fiscal_doc_number", 0)
                for p in r.get("products", []):
                    name = p.get("name", "")
                    price = round(float(p.get("price", 0)), 2)
                    key = (name, price)
                    agg[key]["quantity"] += p.get("quantity", 1)
                    agg[key]["sum"] += p.get("sum", 0)
                    if receipt_id not in agg[key]["receipt_ids"]:
                        agg[key]["cash_sum"] += r.get("cash_sum", 0)
                        agg[key]["card_sum"] += r.get("card_sum", 0)
                    agg[key]["receipt_ids"].add(receipt_id)
            for (name, price), data in agg.items():
                ws3.append([name, data["quantity"], price, data["sum"], data["cash_sum"], data["card_sum"], len(data["receipt_ids"])])
        else:
            ws3.append(["Дата", "ККТ", "ФН", "Чек", "Товар", "Кол-во", "Цена", "Сумма", "Оплата нал", "Оплата карта"])
            for r in receipts:
                if r.get("date", "") < date_from or r.get("date", "") > date_to:
                    continue
                for p in r.get("products", []):
                    ws3.append([
                        r["date"], r.get("kkt_rnm", ""), r.get("fn", ""),
                        r.get("fiscal_doc_number", 0),
                        p.get("name", ""), p.get("quantity", 1), p.get("price", 0),
                        p.get("sum", 0), r.get("cash_sum", 0), r.get("card_sum", 0),
                    ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"ofd_{inn}_{date_from}_{date_to}.xlsx")


@bp.route("/api/export/csv")
def export_csv():
    """Export summary to CSV."""
    from ofd_storage import OfdStorage
    from io import StringIO
    import csv

    inn = request.args.get("inn", "")
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    if not inn:
        return jsonify({"error": "INN required"})

    st = OfdStorage(inn)
    ab = st.load_abonent()

    si = StringIO()
    cw = csv.writer(si)
    cw.writerow(["ККТ", "ФН", "Провайдер", "Смены", "Чеков", "Нал", "Карта", "Возврат", "Сумма", "Ошибки"])
    for k in ab.get("kkt_summary", []):
        t = k.get("totals", {})
        cw.writerow([
            k.get("kkt_name", ""), k.get("fn_current", ""), k.get("provider", ""),
            t.get("shifts", 0), t.get("receipts", {}).get("total", 0),
            t.get("receipts", {}).get("cash", 0), t.get("receipts", {}).get("card", 0),
            t.get("receipts", {}).get("return", 0), t.get("sums", {}).get("total", 0),
            t.get("errors", 0),
        ])

    from flask import Response
    return Response(si.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment;filename=ofd_{inn}_{date_from}_{date_to}.csv"})


@bp.route("/api/export/csv-items")
def export_csv_items():
    """Export receipt items to CSV."""
    from ofd_storage import OfdStorage
    from io import StringIO
    import csv

    inn = request.args.get("inn", "")
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    aggr = request.args.get("aggr", "0") == "1"
    if not inn:
        return jsonify({"error": "INN required"})

    st = OfdStorage(inn)
    items_data = st.load_items()
    receipts = items_data.get("receipts", [])

    si = StringIO()
    cw = csv.writer(si)
    if aggr:
        cw.writerow(["Товар", "Кол-во", "Цена", "Сумма", "Оплата нал", "Оплата карта", "Чеков"])
        from collections import defaultdict
        agg = defaultdict(lambda: {"quantity": 0, "sum": 0.0, "cash_sum": 0.0, "card_sum": 0.0, "receipt_ids": set()})
        for r in receipts:
            if r.get("date", "") < date_from or r.get("date", "") > date_to:
                continue
            receipt_id = r.get("fiscal_doc_number", 0)
            for p in r.get("products", []):
                name = p.get("name", "")
                price = round(float(p.get("price", 0)), 2)
                key = (name, price)
                agg[key]["quantity"] += p.get("quantity", 1)
                agg[key]["sum"] += p.get("sum", 0)
                if receipt_id not in agg[key]["receipt_ids"]:
                    agg[key]["cash_sum"] += r.get("cash_sum", 0)
                    agg[key]["card_sum"] += r.get("card_sum", 0)
                agg[key]["receipt_ids"].add(receipt_id)
        for (name, price), data in agg.items():
            cw.writerow([name, data["quantity"], price, data["sum"], data["cash_sum"], data["card_sum"], len(data["receipt_ids"])])
    else:
        cw.writerow(["Дата", "ККТ", "ФН", "Чек", "Товар", "Кол-во", "Цена", "Сумма", "Оплата нал", "Оплата карта"])
        for r in receipts:
            if r.get("date", "") < date_from or r.get("date", "") > date_to:
                continue
            for p in r.get("products", []):
                cw.writerow([
                    r["date"], r.get("kkt_rnm", ""), r.get("fn", ""),
                    r.get("fiscal_doc_number", 0),
                    p.get("name", ""), p.get("quantity", 1), p.get("price", 0),
                    p.get("sum", 0), r.get("cash_sum", 0), r.get("card_sum", 0),
                ])

    from flask import Response
    return Response(si.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment;filename=ofd_items_{inn}_{date_from}_{date_to}.csv"})


@bp.route("/api/items")
def api_items():
    """Return items data as JSON."""
    from ofd_storage import OfdStorage
    inn = request.args.get("inn", "")
    if not inn:
        return jsonify({"error": "INN required"})
    st = OfdStorage(inn)
    return jsonify(st.load_items())
